from pathlib import Path
import csv
import unicodedata

from django.core.management.base import BaseCommand, CommandError

from ramais_contatos.models import RamalContato
from usuarios.models import Unidade


def limpar(valor):
    if valor is None:
        return ''

    valor = str(valor).strip()

    if valor.endswith('.0'):
        valor = valor[:-2]

    return valor


def normalizar(valor):
    valor = limpar(valor).lower()

    valor = unicodedata.normalize('NFKD', valor)
    valor = ''.join(c for c in valor if not unicodedata.combining(c))

    return valor.strip()


def valor_bool(valor):
    valor = normalizar(valor)

    if valor in ['sim', 's', 'yes', 'y', 'true', '1', 'ativo']:
        return True

    if valor in ['nao', 'não', 'n', 'no', 'false', '0', 'inativo']:
        return False

    return True


def mapear_tipo(valor):
    valor = normalizar(valor)

    mapa = {
        'setor': 'setor',
        'departamento': 'setor',
        'area': 'setor',

        'pessoa': 'pessoa',
        'colaborador': 'pessoa',
        'funcionario': 'pessoa',

        'servico': 'servico',
        'servicos': 'servico',

        'emergencia': 'emergencia',
        'critico': 'emergencia',
        'emergencia / critico': 'emergencia',
    }

    return mapa.get(valor, 'setor')


def detectar_delimitador(primeira_linha):
    if ';' in primeira_linha:
        return ';'

    if ',' in primeira_linha:
        return ','

    return ';'


def ler_csv_com_encoding(caminho, encoding):
    with caminho.open('r', encoding=encoding, newline='') as arquivo:
        primeira_linha = arquivo.readline()
        arquivo.seek(0)

        delimitador = detectar_delimitador(primeira_linha)

        leitor = csv.DictReader(arquivo, delimiter=delimitador)

        for linha in leitor:
            linha_normalizada = {}

            for chave, valor in linha.items():
                chave_normalizada = normalizar(chave)
                linha_normalizada[chave_normalizada] = valor

            yield linha_normalizada


def ler_csv(caminho, encoding_preferencial):
    encodings_para_tentar = []

    if encoding_preferencial:
        encodings_para_tentar.append(encoding_preferencial)

    for encoding in ['utf-8-sig', 'latin-1', 'cp1252']:
        if encoding not in encodings_para_tentar:
            encodings_para_tentar.append(encoding)

    ultimo_erro = None

    for encoding in encodings_para_tentar:
        try:
            linhas = list(ler_csv_com_encoding(caminho, encoding))
            return linhas, encoding
        except UnicodeDecodeError as erro:
            ultimo_erro = erro
            continue

    raise CommandError(
        f'Não foi possível ler o CSV com os encodings testados. Último erro: {ultimo_erro}'
    )


def ler_xlsx(caminho):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CommandError(
            'Para importar XLSX, instale openpyxl no ambiente virtual: pip install openpyxl'
        ) from exc

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))

    if not linhas:
        return []

    cabecalhos = [normalizar(c) for c in linhas[0]]
    resultado = []

    for valores in linhas[1:]:
        linha = {}

        for indice, cabecalho in enumerate(cabecalhos):
            if not cabecalho:
                continue

            valor = valores[indice] if indice < len(valores) else ''
            linha[cabecalho] = valor

        resultado.append(linha)

    return resultado


def obter_valor(linha, *nomes):
    for nome in nomes:
        chave = normalizar(nome)

        if chave in linha:
            return limpar(linha.get(chave))

    return ''


class Command(BaseCommand):
    help = 'Importa ramais e contatos a partir de arquivo CSV ou XLSX.'

    def add_arguments(self, parser):
        parser.add_argument(
            'arquivo',
            type=str,
            help='Caminho do arquivo CSV ou XLSX.'
        )

        parser.add_argument(
            '--encoding',
            type=str,
            default='utf-8-sig',
            help='Encoding para CSV. Padrão: utf-8-sig. O comando também tenta latin-1 e cp1252 automaticamente.'
        )

    def handle(self, *args, **options):
        caminho = Path(options['arquivo'])
        encoding = options['encoding']

        if not caminho.exists():
            raise CommandError(f'Arquivo não encontrado: {caminho}')

        extensao = caminho.suffix.lower()
        encoding_usado = ''

        if extensao == '.csv':
            linhas, encoding_usado = ler_csv(caminho, encoding)
        elif extensao == '.xlsx':
            linhas = ler_xlsx(caminho)
            encoding_usado = 'xlsx'
        else:
            raise CommandError('Formato não suportado. Use .csv ou .xlsx.')

        total_linhas = 0
        total_criados = 0
        total_atualizados = 0
        total_erros = 0

        self.stdout.write(
            self.style.WARNING('Iniciando importação de ramais e contatos...')
        )

        self.stdout.write(
            self.style.WARNING(f'Formato/encoding usado: {encoding_usado}')
        )

        for numero_linha, linha in enumerate(linhas, start=2):
            total_linhas += 1

            unidade_sigla = obter_valor(
                linha,
                'unidade_sigla',
                'sigla_unidade',
                'unidade',
                'codigo_unidade'
            )

            tipo = obter_valor(
                linha,
                'tipo'
            )

            setor = obter_valor(
                linha,
                'setor'
            )

            nome = obter_valor(
                linha,
                'nome',
                'contato',
                'descricao',
                'descrição'
            )

            cargo_funcao = obter_valor(
                linha,
                'cargo_funcao',
                'cargo/função',
                'cargo_funcao',
                'funcao',
                'função',
                'cargo',
                'função/cargo'
            )

            ramal = obter_valor(
                linha,
                'ramal'
            )

            telefone = obter_valor(
                linha,
                'telefone',
                'fone'
            )

            celular = obter_valor(
                linha,
                'celular',
                'whatsapp'
            )

            email = obter_valor(
                linha,
                'email',
                'e-mail',
                'e_mail'
            )

            localizacao = obter_valor(
                linha,
                'localizacao',
                'localização',
                'local'
            )

            observacao = obter_valor(
                linha,
                'observacao',
                'observação',
                'obs'
            )

            ativo_valor = obter_valor(
                linha,
                'ativo'
            )

            ordem_valor = obter_valor(
                linha,
                'ordem'
            )

            if not nome:
                total_erros += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: campo nome é obrigatório.'
                    )
                )
                continue

            unidade = None

            if unidade_sigla:
                unidade = Unidade.objects.filter(
                    sigla__iexact=unidade_sigla
                ).first()

                if not unidade:
                    unidade = Unidade.objects.filter(
                        nome__icontains=unidade_sigla
                    ).first()

                if not unidade:
                    total_erros += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f'Linha {numero_linha}: unidade não encontrada: {unidade_sigla}'
                        )
                    )
                    continue

            tipo_mapeado = mapear_tipo(tipo)

            try:
                ordem = int(ordem_valor) if ordem_valor else 0
            except ValueError:
                ordem = 0

            ativo = valor_bool(ativo_valor)

            contato, criado = RamalContato.objects.update_or_create(
                unidade=unidade,
                nome=nome,
                setor=setor,
                ramal=ramal,
                defaults={
                    'tipo': tipo_mapeado,
                    'cargo_funcao': cargo_funcao,
                    'telefone': telefone,
                    'celular': celular,
                    'email': email,
                    'localizacao': localizacao,
                    'observacao': observacao,
                    'ativo': ativo,
                    'ordem': ordem,
                }
            )

            if criado:
                total_criados += 1
            else:
                total_atualizados += 1

            if (total_criados + total_atualizados) % 100 == 0:
                self.stdout.write(
                    f'Processados {total_criados + total_atualizados} contatos...'
                )

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('Importação finalizada.'))
        self.stdout.write(f'Linhas lidas: {total_linhas}')
        self.stdout.write(f'Contatos criados: {total_criados}')
        self.stdout.write(f'Contatos atualizados: {total_atualizados}')
        self.stdout.write(f'Linhas com erro: {total_erros}')