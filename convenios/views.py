import csv
import os

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from auditoria.models import RegistroAuditoria
from usuarios.models import Unidade
from modulos.models import Modulo
from conteudos.models import ConteudoModulo

from .models import (
    Convenio,
    PlanoConvenio,
    Especialidade,
    RegraAtendimentoConvenio,
    ProcedimentoProibidoPlano,
    ImportacaoMV,
    ItemImportacaoMV,
)


def usuario_eh_admin_ti(user):
    return user.is_superuser or user.groups.filter(name='TI Administrador').exists()


def usuario_pode_acessar_modulo_mv(user):
    if usuario_eh_admin_ti(user):
        return True

    try:
        modulo = Modulo.objects.get(nome='MV / Sistema Hospitalar', ativo=True)
    except Modulo.DoesNotExist:
        return False

    if not modulo.grupos_permitidos.exists():
        return True

    return modulo.grupos_permitidos.filter(
        id__in=user.groups.values_list('id', flat=True)
    ).exists()


def obter_ip_cliente(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')

    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()

    return request.META.get('REMOTE_ADDR')


def registrar_auditoria_importacao(request, importacao, titulo):
    RegistroAuditoria.objects.create(
        modulo='convenios',
        acao='criado',
        titulo=titulo,
        descricao=(
            f'Tipo: {importacao.get_tipo_display()}\n'
            f'Status: {importacao.get_status_display()}\n'
            f'Total linhas: {importacao.total_linhas}\n'
            f'Sucesso: {importacao.total_sucesso}\n'
            f'Erros: {importacao.total_erros}\n'
            f'Mensagem: {importacao.mensagem or "Não informada"}'
        ),
        modelo='ImportacaoMV',
        objeto_id=str(importacao.id),
        usuario=request.user,
        unidade=getattr(request.user, 'unidade', None),
        ip_origem=obter_ip_cliente(request),
    )


def normalizar_texto(valor):
    if valor is None:
        return ''

    return str(valor).strip()


def normalizar_cabecalho(valor):
    valor = normalizar_texto(valor).lower()

    substituicoes = {
        'á': 'a',
        'à': 'a',
        'ã': 'a',
        'â': 'a',
        'é': 'e',
        'ê': 'e',
        'í': 'i',
        'ó': 'o',
        'ô': 'o',
        'õ': 'o',
        'ú': 'u',
        'ç': 'c',
    }

    for antigo, novo in substituicoes.items():
        valor = valor.replace(antigo, novo)

    valor = valor.replace(' ', '_')
    valor = valor.replace('-', '_')
    valor = valor.replace('/', '_')

    while '__' in valor:
        valor = valor.replace('__', '_')

    return valor


def valor_booleano(valor):
    valor = normalizar_texto(valor).lower()

    return valor in [
        'sim',
        's',
        'yes',
        'y',
        '1',
        'true',
        'verdadeiro',
        'ativo',
        'ativa',
    ]


def ler_arquivo_importacao(importacao):
    caminho = importacao.arquivo.path
    extensao = os.path.splitext(caminho)[1].lower()

    if extensao == '.csv':
        return ler_csv(caminho)

    if extensao in ['.xlsx', '.xlsm']:
        return ler_xlsx(caminho)

    raise ValueError('Formato inválido. Envie arquivo CSV ou XLSX.')


def ler_csv(caminho):
    linhas = []

    with open(caminho, 'r', encoding='utf-8-sig', newline='') as arquivo:
        amostra = arquivo.read(2048)
        arquivo.seek(0)

        delimitador = ';'

        if ',' in amostra and ';' not in amostra:
            delimitador = ','

        leitor = csv.DictReader(arquivo, delimiter=delimitador)

        for linha in leitor:
            linhas.append({
                normalizar_cabecalho(chave): normalizar_texto(valor)
                for chave, valor in linha.items()
                if chave is not None
            })

    return linhas


def ler_xlsx(caminho):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError('Biblioteca openpyxl não instalada. Instale com: pip install openpyxl')

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    worksheet = workbook.active

    linhas = []
    cabecalhos = []

    for indice, linha in enumerate(worksheet.iter_rows(values_only=True), start=1):
        valores = list(linha)

        if indice == 1:
            cabecalhos = [normalizar_cabecalho(valor) for valor in valores]
            continue

        if not any(valores):
            continue

        dados = {}

        for posicao, cabecalho in enumerate(cabecalhos):
            if not cabecalho:
                continue

            valor = valores[posicao] if posicao < len(valores) else ''
            dados[cabecalho] = normalizar_texto(valor)

        linhas.append(dados)

    return linhas


def validar_colunas(linhas, colunas_obrigatorias):
    if not linhas:
        raise ValueError('Arquivo vazio ou sem linhas para importar.')

    colunas_arquivo = set(linhas[0].keys())
    faltantes = []

    for coluna in colunas_obrigatorias:
        if coluna not in colunas_arquivo:
            faltantes.append(coluna)

    if faltantes:
        raise ValueError(f'Colunas obrigatórias ausentes: {", ".join(faltantes)}')


def gravar_item_importacao(importacao, linha, status, mensagem, dados):
    ItemImportacaoMV.objects.create(
        importacao=importacao,
        linha=linha,
        status=status,
        mensagem=mensagem,
        dados=dados,
    )


def processar_importacao_convenios_planos(importacao):
    linhas = ler_arquivo_importacao(importacao)

    validar_colunas(
        linhas,
        [
            'codigo_convenio',
            'nome_convenio',
            'nome_plano',
        ]
    )

    total_sucesso = 0
    total_erros = 0

    for indice, dados in enumerate(linhas, start=2):
        try:
            codigo_convenio = normalizar_texto(dados.get('codigo_convenio'))
            nome_convenio = normalizar_texto(dados.get('nome_convenio'))
            tipo_mv = normalizar_texto(dados.get('tipo_mv'))
            codigo_plano = normalizar_texto(dados.get('codigo_plano'))
            nome_plano = normalizar_texto(dados.get('nome_plano'))

            regra_codigo_mv = normalizar_texto(dados.get('regra_codigo_mv'))
            regra_nome_mv = normalizar_texto(dados.get('regra_nome_mv'))
            indice_codigo_mv = normalizar_texto(dados.get('indice_codigo_mv'))
            indice_nome_mv = normalizar_texto(dados.get('indice_nome_mv'))

            if not nome_convenio:
                raise ValueError('Nome do convênio não informado.')

            if not nome_plano:
                raise ValueError('Nome do plano não informado.')

            convenio = None

            if codigo_convenio:
                convenio = Convenio.objects.filter(codigo_mv=codigo_convenio).first()

            if not convenio:
                convenio = Convenio.objects.filter(nome__iexact=nome_convenio).first()

            if not convenio:
                convenio = Convenio.objects.create(
                    codigo_mv=codigo_convenio,
                    nome=nome_convenio,
                    tipo_mv=tipo_mv,
                    ativo=True,
                )
            else:
                convenio.nome = nome_convenio
                convenio.tipo_mv = tipo_mv

                if codigo_convenio:
                    convenio.codigo_mv = codigo_convenio

                convenio.ativo = True
                convenio.save()

            plano = None

            if codigo_plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    codigo_mv=codigo_plano,
                ).first()

            if not plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    nome__iexact=nome_plano,
                ).first()

            if not plano:
                plano = PlanoConvenio.objects.create(
                    convenio=convenio,
                    codigo_mv=codigo_plano,
                    nome=nome_plano,
                    regra_codigo_mv=regra_codigo_mv,
                    regra_nome_mv=regra_nome_mv,
                    indice_codigo_mv=indice_codigo_mv,
                    indice_nome_mv=indice_nome_mv,
                    ativo=True,
                )
            else:
                plano.nome = nome_plano
                plano.codigo_mv = codigo_plano
                plano.regra_codigo_mv = regra_codigo_mv
                plano.regra_nome_mv = regra_nome_mv
                plano.indice_codigo_mv = indice_codigo_mv
                plano.indice_nome_mv = indice_nome_mv
                plano.ativo = True
                plano.save()

            total_sucesso += 1

            gravar_item_importacao(
                importacao,
                indice,
                'sucesso',
                f'Convênio/plano importado: {convenio.nome} - {plano.nome}',
                dados,
            )

        except Exception as erro:
            total_erros += 1

            gravar_item_importacao(
                importacao,
                indice,
                'erro',
                str(erro),
                dados,
            )

    return len(linhas), total_sucesso, total_erros


def processar_importacao_procedimentos_proibidos(importacao):
    linhas = ler_arquivo_importacao(importacao)

    validar_colunas(
        linhas,
        [
            'codigo_convenio',
            'codigo_plano',
            'codigo_procedimento',
            'descricao_procedimento',
        ]
    )

    total_sucesso = 0
    total_erros = 0

    for indice, dados in enumerate(linhas, start=2):
        try:
            codigo_convenio = normalizar_texto(dados.get('codigo_convenio'))
            nome_convenio = normalizar_texto(dados.get('nome_convenio'))
            codigo_plano = normalizar_texto(dados.get('codigo_plano'))
            nome_plano = normalizar_texto(dados.get('nome_plano'))
            codigo_procedimento = normalizar_texto(dados.get('codigo_procedimento'))
            descricao_procedimento = normalizar_texto(dados.get('descricao_procedimento'))

            if not codigo_procedimento:
                raise ValueError('Código do procedimento não informado.')

            if not descricao_procedimento:
                raise ValueError('Descrição do procedimento não informada.')

            convenio = None

            if codigo_convenio:
                convenio = Convenio.objects.filter(codigo_mv=codigo_convenio).first()

            if not convenio and nome_convenio:
                convenio = Convenio.objects.filter(nome__iexact=nome_convenio).first()

            if not convenio:
                raise ValueError('Convênio não localizado. Importe convênios e planos antes.')

            plano = None

            if codigo_plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    codigo_mv=codigo_plano,
                ).first()

            if not plano and nome_plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    nome__iexact=nome_plano,
                ).first()

            if not plano:
                raise ValueError('Plano não localizado. Importe convênios e planos antes.')

            procedimento, criado = ProcedimentoProibidoPlano.objects.update_or_create(
                plano=plano,
                codigo_procedimento=codigo_procedimento,
                defaults={
                    'convenio': convenio,
                    'descricao_procedimento': descricao_procedimento,
                    'ativo': True,
                }
            )

            total_sucesso += 1

            acao = 'criado' if criado else 'atualizado'

            gravar_item_importacao(
                importacao,
                indice,
                'sucesso',
                f'Procedimento {acao}: {procedimento.codigo_procedimento} - {procedimento.descricao_procedimento}',
                dados,
            )

        except Exception as erro:
            total_erros += 1

            gravar_item_importacao(
                importacao,
                indice,
                'erro',
                str(erro),
                dados,
            )

    return len(linhas), total_sucesso, total_erros


def processar_importacao_regras_atendimento(importacao):
    linhas = ler_arquivo_importacao(importacao)

    validar_colunas(
        linhas,
        [
            'unidade',
            'codigo_convenio',
            'codigo_plano',
            'tipo_atendimento',
            'status',
        ]
    )

    total_sucesso = 0
    total_erros = 0

    for indice, dados in enumerate(linhas, start=2):
        try:
            unidade_valor = normalizar_texto(dados.get('unidade'))
            codigo_convenio = normalizar_texto(dados.get('codigo_convenio'))
            nome_convenio = normalizar_texto(dados.get('nome_convenio'))
            codigo_plano = normalizar_texto(dados.get('codigo_plano'))
            nome_plano = normalizar_texto(dados.get('nome_plano'))
            tipo_atendimento = normalizar_texto(dados.get('tipo_atendimento')).lower()
            especialidade_nome = normalizar_texto(dados.get('especialidade'))
            status = normalizar_texto(dados.get('status')).lower()
            exige_autorizacao = valor_booleano(dados.get('exige_autorizacao'))
            observacao = normalizar_texto(dados.get('observacao'))

            unidade = Unidade.objects.filter(sigla__iexact=unidade_valor).first()

            if not unidade:
                unidade = Unidade.objects.filter(nome__iexact=unidade_valor).first()

            if not unidade:
                raise ValueError('Unidade não localizada.')

            convenio = None

            if codigo_convenio:
                convenio = Convenio.objects.filter(codigo_mv=codigo_convenio).first()

            if not convenio and nome_convenio:
                convenio = Convenio.objects.filter(nome__iexact=nome_convenio).first()

            if not convenio:
                raise ValueError('Convênio não localizado.')

            plano = None

            if codigo_plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    codigo_mv=codigo_plano,
                ).first()

            if not plano and nome_plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    nome__iexact=nome_plano,
                ).first()

            if not plano:
                raise ValueError('Plano não localizado.')

            tipos_validos = [item[0] for item in RegraAtendimentoConvenio.TIPO_ATENDIMENTO_CHOICES]

            if tipo_atendimento not in tipos_validos:
                raise ValueError(f'Tipo de atendimento inválido: {tipo_atendimento}')

            status_validos = [item[0] for item in RegraAtendimentoConvenio.STATUS_CHOICES]

            if status not in status_validos:
                raise ValueError(f'Status inválido: {status}')

            especialidade = None

            if especialidade_nome:
                especialidade, _ = Especialidade.objects.get_or_create(
                    nome=especialidade_nome,
                    defaults={
                        'ativo': True,
                    }
                )

            regra, criado = RegraAtendimentoConvenio.objects.update_or_create(
                unidade=unidade,
                convenio=convenio,
                plano=plano,
                tipo_atendimento=tipo_atendimento,
                especialidade=especialidade,
                defaults={
                    'status': status,
                    'exige_autorizacao': exige_autorizacao,
                    'observacao': observacao,
                    'ativo': True,
                }
            )

            total_sucesso += 1

            acao = 'criada' if criado else 'atualizada'

            gravar_item_importacao(
                importacao,
                indice,
                'sucesso',
                f'Regra {acao}: {regra}',
                dados,
            )

        except Exception as erro:
            total_erros += 1

            gravar_item_importacao(
                importacao,
                indice,
                'erro',
                str(erro),
                dados,
            )

    return len(linhas), total_sucesso, total_erros


def processar_importacao_mv(importacao):
    importacao.status = 'processando'
    importacao.iniciado_em = timezone.now()
    importacao.mensagem = ''
    importacao.save()

    try:
        with transaction.atomic():
            if importacao.tipo == 'convenios_planos':
                total_linhas, total_sucesso, total_erros = processar_importacao_convenios_planos(importacao)

            elif importacao.tipo == 'procedimentos_proibidos':
                total_linhas, total_sucesso, total_erros = processar_importacao_procedimentos_proibidos(importacao)

            elif importacao.tipo == 'regras_atendimento':
                total_linhas, total_sucesso, total_erros = processar_importacao_regras_atendimento(importacao)

            else:
                raise ValueError('Tipo de importação inválido.')

        importacao.total_linhas = total_linhas
        importacao.total_sucesso = total_sucesso
        importacao.total_erros = total_erros
        importacao.finalizado_em = timezone.now()

        if total_erros > 0 and total_sucesso > 0:
            importacao.status = 'concluida_com_erros'
            importacao.mensagem = 'Importação concluída com alguns erros.'
        elif total_erros > 0 and total_sucesso == 0:
            importacao.status = 'erro'
            importacao.mensagem = 'Importação finalizada com erros. Nenhuma linha foi importada.'
        else:
            importacao.status = 'concluida'
            importacao.mensagem = 'Importação concluída com sucesso.'

        importacao.save()

    except Exception as erro:
        importacao.status = 'erro'
        importacao.mensagem = str(erro)
        importacao.finalizado_em = timezone.now()
        importacao.save()

        gravar_item_importacao(
            importacao,
            0,
            'erro',
            str(erro),
            {},
        )


def buscar_conteudos_mv_por_tipo(request, tipo):
    modulo = get_object_or_404(
        Modulo,
        nome='MV / Sistema Hospitalar',
        ativo=True
    )

    unidade_usuario = getattr(request.user, 'unidade', None)

    conteudos = ConteudoModulo.objects.filter(
        modulo=modulo,
        tipo=tipo,
        ativo=True
    ).filter(
        Q(unidade=unidade_usuario) |
        Q(unidade__isnull=True)
    ).select_related(
        'unidade',
        'modulo'
    ).prefetch_related(
        'grupos_permitidos'
    )

    if not usuario_eh_admin_ti(request.user):
        grupos_usuario = request.user.groups.all()

        conteudos = conteudos.filter(
            Q(grupos_permitidos__in=grupos_usuario) |
            Q(grupos_permitidos__isnull=True)
        ).distinct()

    busca = request.GET.get('busca', '').strip()
    unidade_id = request.GET.get('unidade', '').strip()

    if busca:
        conteudos = conteudos.filter(
            Q(titulo__icontains=busca) |
            Q(descricao__icontains=busca) |
            Q(link_externo__icontains=busca) |
            Q(arquivo__icontains=busca)
        )

    if unidade_id:
        if unidade_id == 'geral':
            conteudos = conteudos.filter(unidade__isnull=True)
        else:
            conteudos = conteudos.filter(unidade_id=unidade_id)

    return conteudos.order_by('ordem', 'titulo'), modulo


def renderizar_conteudo_mv_tipo(request, tipo, titulo, subtitulo, icone, etiqueta):
    if not usuario_pode_acessar_modulo_mv(request.user):
        return render(request, 'core/sem_permissao.html', status=403)

    conteudos, modulo = buscar_conteudos_mv_por_tipo(request, tipo)

    unidades = Unidade.objects.filter(
        ativo=True
    ).order_by(
        'nome'
    )

    return render(request, 'convenios/mv_conteudos_tipo.html', {
        'modulo': modulo,
        'conteudos': conteudos,
        'unidades': unidades,
        'busca': request.GET.get('busca', '').strip(),
        'unidade_id': request.GET.get('unidade', '').strip(),
        'titulo': titulo,
        'subtitulo': subtitulo,
        'icone': icone,
        'etiqueta': etiqueta,
        'tipo': tipo,
        'total_conteudos': conteudos.count(),
    })


@login_required(login_url='/')
def mv_contingencia(request):
    return renderizar_conteudo_mv_tipo(
        request=request,
        tipo='contingencia',
        titulo='Contingência MV',
        subtitulo='Orientações para indisponibilidade, lentidão ou falha do sistema MV.',
        icone='🚨',
        etiqueta='Contingência',
    )


@login_required(login_url='/')
def mv_chamados(request):
    return renderizar_conteudo_mv_tipo(
        request=request,
        tipo='chamado',
        titulo='Chamados MV',
        subtitulo='Links, orientações e procedimentos para abertura e acompanhamento de chamados junto à MV.',
        icone='🧾',
        etiqueta='Chamado MV',
    )


@login_required(login_url='/')
def mv_links(request):
    return renderizar_conteudo_mv_tipo(
        request=request,
        tipo='link',
        titulo='Links úteis MV',
        subtitulo='Acessos rápidos relacionados ao sistema MV, portais, fornecedores e ambientes internos.',
        icone='🔗',
        etiqueta='Link útil',
    )


@login_required(login_url='/')
def mv_observacoes(request):
    return renderizar_conteudo_mv_tipo(
        request=request,
        tipo='observacao',
        titulo='Observações MV',
        subtitulo='Orientações internas, alertas operacionais e observações relacionadas ao uso do sistema MV.',
        icone='📝',
        etiqueta='Observação',
    )


@login_required(login_url='/')
def importacoes_mv(request):
    if not usuario_eh_admin_ti(request.user):
        return render(request, 'core/sem_permissao.html', status=403)

    importacoes = ImportacaoMV.objects.select_related(
        'usuario'
    ).order_by(
        '-criado_em'
    )[:100]

    return render(request, 'convenios/importacoes_mv.html', {
        'importacoes': importacoes,
        'tipos_importacao': ImportacaoMV.TIPO_CHOICES,
    })


@login_required(login_url='/')
def nova_importacao_mv(request):
    if not usuario_eh_admin_ti(request.user):
        return render(request, 'core/sem_permissao.html', status=403)

    if request.method == 'POST':
        tipo = request.POST.get('tipo', '').strip()
        arquivo = request.FILES.get('arquivo')

        erros = []

        tipos_validos = [item[0] for item in ImportacaoMV.TIPO_CHOICES]

        if tipo not in tipos_validos:
            erros.append('Tipo de importação inválido.')

        if not arquivo:
            erros.append('Selecione um arquivo CSV ou XLSX.')

        if arquivo:
            extensao = os.path.splitext(arquivo.name)[1].lower()

            if extensao not in ['.csv', '.xlsx', '.xlsm']:
                erros.append('Formato inválido. Envie arquivo CSV ou XLSX.')

        if erros:
            for erro in erros:
                messages.error(request, erro)

            return redirect('/portal/modulos/mv/importacoes/nova/')

        importacao = ImportacaoMV.objects.create(
            tipo=tipo,
            arquivo=arquivo,
            usuario=request.user,
            status='pendente',
        )

        processar_importacao_mv(importacao)

        registrar_auditoria_importacao(
            request,
            importacao,
            f'Importação MV realizada: {importacao.get_tipo_display()}'
        )

        if importacao.status == 'concluida':
            messages.success(request, 'Importação concluída com sucesso.')
        elif importacao.status == 'concluida_com_erros':
            messages.error(request, 'Importação concluída com erros. Verifique o log.')
        else:
            messages.error(request, 'Importação finalizada com erro. Verifique o log.')

        return redirect(f'/portal/modulos/mv/importacoes/{importacao.id}/')

    return render(request, 'convenios/nova_importacao_mv.html', {
        'tipos_importacao': ImportacaoMV.TIPO_CHOICES,
    })


@login_required(login_url='/')
def detalhe_importacao_mv(request, importacao_id):
    if not usuario_eh_admin_ti(request.user):
        return render(request, 'core/sem_permissao.html', status=403)

    importacao = get_object_or_404(
        ImportacaoMV.objects.select_related('usuario'),
        id=importacao_id
    )

    itens = importacao.itens.all().order_by('linha')

    return render(request, 'convenios/detalhe_importacao_mv.html', {
        'importacao': importacao,
        'itens': itens,
    })


@login_required(login_url='/')
def baixar_modelo_importacao_mv(request, tipo):
    if not usuario_eh_admin_ti(request.user):
        return render(request, 'core/sem_permissao.html', status=403)

    modelos = {
        'convenios_planos': [
            'codigo_convenio',
            'nome_convenio',
            'tipo_mv',
            'codigo_plano',
            'nome_plano',
            'regra_codigo_mv',
            'regra_nome_mv',
            'indice_codigo_mv',
            'indice_nome_mv',
        ],
        'procedimentos_proibidos': [
            'codigo_convenio',
            'nome_convenio',
            'codigo_plano',
            'nome_plano',
            'codigo_procedimento',
            'descricao_procedimento',
        ],
        'regras_atendimento': [
            'unidade',
            'codigo_convenio',
            'nome_convenio',
            'codigo_plano',
            'nome_plano',
            'tipo_atendimento',
            'especialidade',
            'status',
            'exige_autorizacao',
            'observacao',
        ],
    }

    if tipo not in modelos:
        messages.error(request, 'Modelo de importação inválido.')
        return redirect('/portal/modulos/mv/importacoes/')

    nome_arquivo = f'modelo_{tipo}.csv'

    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

    response.write('\ufeff')

    writer = csv.writer(
        response,
        delimiter=';',
        quotechar='"',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator='\n'
    )

    writer.writerow(modelos[tipo])

    if tipo == 'convenios_planos':
        writer.writerow([
            '001',
            'BRADESCO SAUDE',
            'CONVENIO',
            'P001',
            'PLANO BASICO',
            'R001',
            'REGRA BASICA',
            'I001',
            'INDICE PADRAO',
        ])

    elif tipo == 'procedimentos_proibidos':
        writer.writerow([
            '001',
            'BRADESCO SAUDE',
            'P001',
            'PLANO BASICO',
            '40302784',
            'VITAMINA B1',
        ])

    elif tipo == 'regras_atendimento':
        writer.writerow([
            'HSFOS',
            '001',
            'BRADESCO SAUDE',
            'P001',
            'PLANO BASICO',
            'consulta',
            'Ortopedia',
            'consultar_autorizacao',
            'sim',
            'Necessário validar autorização antes do atendimento.',
        ])

    return response