from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from usuarios.models import Unidade
from convenios.models import (
    Convenio,
    PlanoConvenio,
    Especialidade,
    RegraAtendimentoConvenio,
)


def limpar(valor):
    if valor is None:
        return ''

    valor = str(valor).strip()

    if valor.endswith('.0'):
        valor = valor[:-2]

    return valor


def normalizar(valor):
    valor = limpar(valor).lower()

    valor = unicodedata.normalize('NFKD', valor)
    valor = ''.join([c for c in valor if not unicodedata.combining(c)])

    valor = valor.replace('ç', 'c')
    valor = valor.replace('ã', 'a')
    valor = valor.replace('õ', 'o')
    valor = valor.replace('á', 'a')
    valor = valor.replace('é', 'e')
    valor = valor.replace('í', 'i')
    valor = valor.replace('ó', 'o')
    valor = valor.replace('ú', 'u')

    valor = valor.replace('-', '_')
    valor = valor.replace(' ', '_')
    valor = valor.replace('/', '_')

    while '__' in valor:
        valor = valor.replace('__', '_')

    return valor.strip('_')


def texto_sim_nao_para_booleano(valor):
    valor = normalizar(valor)

    if valor in ['sim', 's', 'yes', 'y', 'true', '1']:
        return True

    return False


def normalizar_tipo_atendimento(valor):
    valor = normalizar(valor)

    mapa = {
        'consulta': 'consulta',
        'consultas': 'consulta',

        'pronto_atendimento': 'pronto_atendimento',
        'pa': 'pronto_atendimento',
        'ps': 'pronto_atendimento',
        'pronto_socorro': 'pronto_atendimento',

        'exame': 'exame',
        'exames': 'exame',

        'internacao': 'internacao',
        'internacao_hospitalar': 'internacao',

        'cirurgia': 'cirurgia',
        'cirurgias': 'cirurgia',

        'terapia': 'terapia',
        'terapias': 'terapia',

        'pediatria': 'pediatria',
        'consulta_pediatria': 'pediatria',
    }

    return mapa.get(valor, valor)


def normalizar_status(valor):
    valor = normalizar(valor)

    mapa = {
        'aceito': 'aceito',
        'aceita': 'aceito',
        'sim': 'aceito',
        's': 'aceito',
        'permitido': 'aceito',
        'liberado': 'aceito',

        'nao_aceito': 'nao_aceito',
        'nao_aceita': 'nao_aceito',
        'nao': 'nao_aceito',
        'n': 'nao_aceito',
        'negado': 'nao_aceito',
        'bloqueado': 'nao_aceito',

        'consultar_autorizacao': 'consultar_autorizacao',
        'autorizacao': 'consultar_autorizacao',
        'consulta_autorizacao': 'consultar_autorizacao',
        'necessita_autorizacao': 'consultar_autorizacao',
        'exige_autorizacao': 'consultar_autorizacao',

        'suspenso': 'suspenso',
        'suspenso_temporariamente': 'suspenso',
        'temporariamente_suspenso': 'suspenso',
    }

    return mapa.get(valor, valor)


def localizar_colunas(cabecalho):
    colunas = {}

    for indice, nome_coluna in enumerate(cabecalho):
        nome = normalizar(nome_coluna)

        if nome:
            colunas[nome] = indice

    return colunas


def valor_linha(row, colunas, nome_coluna):
    indice = colunas.get(nome_coluna)

    if indice is None:
        return ''

    if indice >= len(row):
        return ''

    return limpar(row[indice])


class Command(BaseCommand):
    help = 'Importa regras de atendimento de convênios por unidade, plano, tipo e especialidade.'

    def add_arguments(self, parser):
        parser.add_argument(
            'arquivo',
            type=str,
            help='Caminho da planilha regras_convenios.xlsx'
        )

    def handle(self, *args, **options):
        caminho = Path(options['arquivo'])

        if not caminho.exists():
            raise CommandError(f'Arquivo não encontrado: {caminho}')

        workbook = load_workbook(caminho, data_only=True)
        sheet = workbook.active

        cabecalho = [cell.value for cell in sheet[1]]
        colunas = localizar_colunas(cabecalho)

        colunas_obrigatorias = [
            'unidade_sigla',
            'convenio',
            'plano',
            'tipo_atendimento',
            'status',
        ]

        for coluna in colunas_obrigatorias:
            if coluna not in colunas:
                raise CommandError(
                    f'Coluna obrigatória não encontrada: {coluna}'
                )

        total_criado = 0
        total_atualizado = 0
        total_erro = 0

        tipos_validos = [
            codigo for codigo, nome in RegraAtendimentoConvenio.TIPO_ATENDIMENTO_CHOICES
        ]

        status_validos = [
            codigo for codigo, nome in RegraAtendimentoConvenio.STATUS_CHOICES
        ]

        for numero_linha, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            unidade_sigla = valor_linha(row, colunas, 'unidade_sigla')
            convenio_texto = valor_linha(row, colunas, 'convenio')
            plano_texto = valor_linha(row, colunas, 'plano')
            tipo_atendimento = normalizar_tipo_atendimento(
                valor_linha(row, colunas, 'tipo_atendimento')
            )
            especialidade_texto = valor_linha(row, colunas, 'especialidade')
            status = normalizar_status(
                valor_linha(row, colunas, 'status')
            )
            exige_autorizacao = texto_sim_nao_para_booleano(
                valor_linha(row, colunas, 'exige_autorizacao')
            )
            observacao = valor_linha(row, colunas, 'observacao')

            if not unidade_sigla and not convenio_texto and not plano_texto:
                continue

            try:
                unidade = Unidade.objects.get(sigla__iexact=unidade_sigla)
            except Unidade.DoesNotExist:
                total_erro += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: unidade não encontrada: {unidade_sigla}'
                    )
                )
                continue

            convenio = Convenio.objects.filter(
                nome__iexact=convenio_texto
            ).first()

            if not convenio:
                convenio = Convenio.objects.filter(
                    codigo_mv=convenio_texto
                ).first()

            if not convenio:
                total_erro += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: convênio não encontrado: {convenio_texto}'
                    )
                )
                continue

            plano = PlanoConvenio.objects.filter(
                convenio=convenio,
                nome__iexact=plano_texto
            ).first()

            if not plano:
                plano = PlanoConvenio.objects.filter(
                    convenio=convenio,
                    codigo_mv=plano_texto
                ).first()

            if not plano:
                total_erro += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: plano não encontrado: {convenio.nome} - {plano_texto}'
                    )
                )
                continue

            if tipo_atendimento not in tipos_validos:
                total_erro += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: tipo de atendimento inválido: {tipo_atendimento}'
                    )
                )
                continue

            if status not in status_validos:
                total_erro += 1
                self.stdout.write(
                    self.style.ERROR(
                        f'Linha {numero_linha}: status inválido: {status}'
                    )
                )
                continue

            especialidade = None

            if especialidade_texto:
                especialidade, _ = Especialidade.objects.get_or_create(
                    nome=especialidade_texto,
                    defaults={
                        'ativo': True,
                    }
                )

            regra, criado = RegraAtendimentoConvenio.objects.update_or_create(
                unidade=unidade,
                convenio=convenio,
                plano=plano,
                tipo_atendimento=tipo_atendimento,
                especialidade=especialidade,
                defaults={
                    'status': status,
                    'exige_autorizacao': exige_autorizacao,
                    'observacao': observacao,
                    'ativo': True,
                }
            )

            if criado:
                total_criado += 1
                self.stdout.write(
                    self.style.SUCCESS(
                        f'Linha {numero_linha}: regra criada - {regra}'
                    )
                )
            else:
                total_atualizado += 1
                self.stdout.write(
                    f'Linha {numero_linha}: regra atualizada - {regra}'
                )

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('Importação finalizada.'))
        self.stdout.write(f'Regras criadas: {total_criado}')
        self.stdout.write(f'Regras atualizadas: {total_atualizado}')
        self.stdout.write(f'Linhas com erro: {total_erro}')