from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from convenios.models import Convenio, PlanoConvenio


def limpar(valor):
    if valor is None:
        return ''

    valor = str(valor).strip()

    if valor.endswith('.0'):
        valor = valor[:-2]

    return valor


def normalizar_label(valor):
    return limpar(valor).replace(':', '').strip().lower()


class Command(BaseCommand):
    help = 'Importa convênios e planos a partir dos relatórios do MV.'

    def add_arguments(self, parser):
        parser.add_argument(
            'arquivo_convenios',
            type=str,
            help='Caminho do arquivo r_convenio.xlsx'
        )

        parser.add_argument(
            'arquivo_planos',
            type=str,
            help='Caminho do arquivo r_conpla.xlsx'
        )

    def handle(self, *args, **options):
        arquivo_convenios = Path(options['arquivo_convenios'])
        arquivo_planos = Path(options['arquivo_planos'])

        if not arquivo_convenios.exists():
            raise CommandError(f'Arquivo não encontrado: {arquivo_convenios}')

        if not arquivo_planos.exists():
            raise CommandError(f'Arquivo não encontrado: {arquivo_planos}')

        self.stdout.write(self.style.WARNING('Importando lista de convênios...'))
        total_convenios = self.importar_convenios(arquivo_convenios)

        self.stdout.write(self.style.WARNING('Importando planos dos convênios...'))
        total_planos = self.importar_planos(arquivo_planos)

        self.stdout.write(self.style.SUCCESS('Importação finalizada.'))
        self.stdout.write(f'Convênios processados: {total_convenios}')
        self.stdout.write(f'Planos processados: {total_planos}')

        self.stdout.write(self.style.WARNING(
            'Atenção: este relatório importa convênios e planos. '
            'As regras de aceito/não aceito por especialidade ainda devem ser cadastradas em Regras de atendimento.'
        ))

    def importar_convenios(self, caminho):
        workbook = load_workbook(caminho, data_only=True)
        sheet = workbook.active

        total = 0

        for row in sheet.iter_rows(min_row=1, values_only=True):
            codigo = limpar(row[1]) if len(row) > 1 else ''
            nome = limpar(row[4]) if len(row) > 4 else ''
            tipo = limpar(row[7]) if len(row) > 7 else ''

            if not codigo or not nome:
                continue

            convenio, criado = Convenio.objects.get_or_create(
                nome=nome,
                defaults={
                    'codigo_mv': codigo,
                    'tipo_mv': tipo,
                    'ativo': True,
                }
            )

            convenio.codigo_mv = codigo
            convenio.tipo_mv = tipo
            convenio.ativo = True
            convenio.save()

            total += 1

            if criado:
                self.stdout.write(self.style.SUCCESS(f'Convênio criado: {codigo} - {nome}'))
            else:
                self.stdout.write(f'Convênio atualizado: {codigo} - {nome}')

        return total

    def importar_planos(self, caminho):
        workbook = load_workbook(caminho, data_only=True)
        sheet = workbook.active

        total = 0
        convenio_atual = None

        for linha in range(1, sheet.max_row + 1):
            col_e = normalizar_label(sheet.cell(linha, 5).value)
            col_f = normalizar_label(sheet.cell(linha, 6).value)
            col_g = normalizar_label(sheet.cell(linha, 7).value)

            if col_e == 'convênio' or col_e == 'convenio':
                codigo_convenio = limpar(sheet.cell(linha, 9).value)
                nome_convenio = limpar(sheet.cell(linha, 13).value)

                if nome_convenio:
                    convenio_atual, _ = Convenio.objects.get_or_create(
                        nome=nome_convenio,
                        defaults={
                            'codigo_mv': codigo_convenio,
                            'ativo': True,
                        }
                    )

                    if codigo_convenio:
                        convenio_atual.codigo_mv = codigo_convenio

                    convenio_atual.ativo = True
                    convenio_atual.save()

                continue

            if col_g == 'tipo' and convenio_atual:
                tipo_mv = limpar(sheet.cell(linha, 13).value)

                if tipo_mv:
                    convenio_atual.tipo_mv = tipo_mv
                    convenio_atual.save()

                continue

            if col_f == 'plano' and convenio_atual:
                codigo_plano = limpar(sheet.cell(linha, 11).value)
                nome_plano = limpar(sheet.cell(linha, 13).value)

                if not nome_plano:
                    continue

                regra_codigo = ''
                regra_nome = ''
                indice_codigo = ''
                indice_nome = ''

                # O relatório do MV costuma trazer:
                # linha do Plano
                # linha em branco
                # linha Regra
                # linha Índice
                linha_regra = linha + 2
                linha_indice = linha + 3

                if linha_regra <= sheet.max_row:
                    label_regra = normalizar_label(sheet.cell(linha_regra, 6).value)

                    if label_regra == 'regra':
                        regra_codigo = limpar(sheet.cell(linha_regra, 11).value)
                        regra_nome = limpar(sheet.cell(linha_regra, 13).value)

                if linha_indice <= sheet.max_row:
                    label_indice = normalizar_label(sheet.cell(linha_indice, 6).value)

                    if label_indice == 'índice' or label_indice == 'indice':
                        indice_codigo = limpar(sheet.cell(linha_indice, 11).value)
                        indice_nome = limpar(sheet.cell(linha_indice, 13).value)

                if codigo_plano:
                    plano = PlanoConvenio.objects.filter(
                        convenio=convenio_atual,
                        codigo_mv=codigo_plano
                    ).first()

                    if not plano:
                        plano = PlanoConvenio(
                            convenio=convenio_atual,
                            codigo_mv=codigo_plano,
                        )
                else:
                    plano = PlanoConvenio.objects.filter(
                        convenio=convenio_atual,
                        nome=nome_plano,
                        codigo_mv=''
                    ).first()

                    if not plano:
                        plano = PlanoConvenio(
                            convenio=convenio_atual,
                            nome=nome_plano,
                        )

                plano.nome = nome_plano
                plano.regra_codigo_mv = regra_codigo
                plano.regra_nome_mv = regra_nome
                plano.indice_codigo_mv = indice_codigo
                plano.indice_nome_mv = indice_nome
                plano.ativo = True
                plano.save()

                total += 1

                self.stdout.write(
                    f'Plano importado: {convenio_atual.nome} - {nome_plano} [{codigo_plano}]'
                )

        return total